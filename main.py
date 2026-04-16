"""
Azure Automation Python 3 Runbook
Application Gateway SSL Certificate Expiry Notification

What it does:
  - Scans ALL Application Gateways across one or more subscriptions (read-only)
  - Parses each SSL certificate's expiry from the PKCS#7 publicCertData field
  - Builds an Excel workbook (all certs) and an HTML email (top-10 soonest expiring)
  - Sends both via Azure Communication Services (ACS) Email

Config resolution order for every setting:
  1. Azure Automation variable / credential asset
  2. Environment variable  (local .env via python-dotenv)
  3. Hard-coded default or sys.exit(1) for required values

All Azure operations are READ-ONLY (list + get).  No resources are created,
modified, or deleted.
"""

# =============================================================================
# Standard-library imports
# =============================================================================
import base64
import datetime
import html as _html
import io
import logging
import os
import sys
from typing import Any, Dict, List, Optional

# =============================================================================
# Automation runtime detection + local .env loading
# =============================================================================
# automationassets is only available inside Azure Automation.
# When running locally the ImportError is caught and python-dotenv is used
# to load a .env file instead.
try:
    import automationassets  # type: ignore
    _IN_AUTOMATION = True
except ImportError:
    _IN_AUTOMATION = False
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass  # python-dotenv is optional; env vars may already be set

# =============================================================================
# Logging
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(name)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
# Suppress verbose Azure SDK HTTP request/response dumps (kept at WARNING+)
logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)
logging.getLogger("azure.identity").setLevel(logging.WARNING)

logger = logging.getLogger("main")

# =============================================================================
# Third-party imports  (must be installed as Automation Account packages)
# =============================================================================
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from azure.identity import ClientSecretCredential
from azure.mgmt.network import NetworkManagementClient
from azure.mgmt.subscription import SubscriptionClient
from cryptography import x509
from cryptography.hazmat.primitives.serialization import pkcs7 as _pkcs7
from cryptography.hazmat.primitives.serialization import Encoding as _Encoding


# =============================================================================
# SECTION 1 — Configuration
# =============================================================================

def _auto_var(name: str) -> str:
    """Read a string variable from the Automation Account. Returns '' on failure."""
    try:
        return str(automationassets.get_automation_variable(name) or "")
    except Exception:
        return ""


def _auto_cred(name: str) -> tuple:
    """Read a credential from the Automation Account. Returns (username, password)."""
    try:
        cred = automationassets.get_automation_credential(name)
        return cred.get("username", ""), cred.get("password", "")
    except Exception:
        return "", ""


def _require(value: str, setting_name: str) -> str:
    """Exit with an error if a required config value is empty."""
    value = value.strip()
    if not value:
        logger.error("Required setting '%s' is missing. Exiting.", setting_name)
        sys.exit(1)
    return value


def _load_config() -> dict:
    """
    Resolve all configuration values.

    Automation Account assets take priority; environment variables are the
    fallback for local runs.  Required values that are missing cause sys.exit(1).
    """
    # -- Credentials ----------------------------------------------------------
    if _IN_AUTOMATION:
        # Service-principal client ID + secret stored as an Automation credential
        client_id, client_secret = _auto_cred("AzureSPCredential")
        tenant_id = _auto_var("AZURE_TENANT_ID") or os.getenv("AZURE_TENANT_ID", "")
    else:
        client_id     = os.getenv("AZURE_CLIENT_ID", "")
        client_secret = os.getenv("AZURE_CLIENT_SECRET", "")
        tenant_id     = os.getenv("AZURE_TENANT_ID", "")

    def _resolve(auto_var_name: str, env_var: str, default: str = "") -> str:
        """Try Automation variable first, then env var, then default."""
        if _IN_AUTOMATION:
            val = _auto_var(auto_var_name)
            if val:
                return val
        return os.getenv(env_var, default)

    # DRY_RUN skips the ACS call and writes files locally instead
    dry_run = os.getenv("DRY_RUN", "").strip().lower() in ("1", "true", "yes")

    acs_connection_string = _resolve("ACS_CONNECTION_STRING", "ACS_CONNECTION_STRING")
    acs_sender            = _resolve("ACS_SENDER_ADDRESS",    "ACS_SENDER_ADDRESS")
    notify_to             = _resolve("AG_NOTIFICATION_TO",    "AG_NOTIFICATION_TO")

    # ACS settings and recipients are only required when actually sending
    if not dry_run:
        acs_connection_string = _require(acs_connection_string, "ACS_CONNECTION_STRING")
        acs_sender            = _require(acs_sender,            "ACS_SENDER_ADDRESS")
        notify_to             = _require(notify_to,             "AG_NOTIFICATION_TO")

    return {
        "client_id":        _require(client_id,     "AZURE_CLIENT_ID / AzureSPCredential.username"),
        "client_secret":    _require(client_secret, "AZURE_CLIENT_SECRET / AzureSPCredential.password"),
        "tenant_id":        _require(tenant_id,     "AZURE_TENANT_ID / AG_SP_TENANT_ID"),
        "subscription_ids": _require(
            _resolve("AG_SUBSCRIPTION_IDS", "AG_SUBSCRIPTION_IDS"),
            "AG_SUBSCRIPTION_IDS",
        ),
        # How many days before expiry counts as a warning (default 14)
        "alert_days":             int(_resolve("AG_ALERT_DAYS", "AG_ALERT_DAYS", "14")),
        "acs_connection_string":  acs_connection_string,
        "acs_sender":             acs_sender,
        "notify_to":              notify_to,
        "dry_run":                dry_run,
    }


# =============================================================================
# SECTION 2 — Certificate Scanning  (read-only Azure API calls)
# =============================================================================

def _leaf_cert(chain: list) -> "x509.Certificate":
    """
    Return the end-entity (leaf) certificate from a PKCS#7 chain.

    Strategy: a leaf cert's subject does not appear as the issuer of any
    other cert in the chain.  Among candidates pick the one expiring soonest
    (shortest lifetime == most likely to be the end-entity cert).
    """
    issuer_keys = {c.issuer.public_bytes(_Encoding.DER) for c in chain}
    leaves      = [c for c in chain if c.subject.public_bytes(_Encoding.DER) not in issuer_keys]
    candidates  = leaves or chain  # fall back to full chain if topology unclear

    def _expiry(c: "x509.Certificate") -> datetime.datetime:
        try:
            return c.not_valid_after_utc           # cryptography >= 42
        except AttributeError:
            return c.not_valid_after.replace(tzinfo=datetime.timezone.utc)

    return min(candidates, key=_expiry)


def _parse_expiry(public_cert_data_b64: str):
    """
    Parse the Azure publicCertData field and return (subject, issuer, expiry_utc).

    Azure wraps uploaded certificates in a PKCS#7 SignedData (.p7b) container
    and base64-encodes the result.  Four decode attempts are made in order:

      1. base64 → PKCS#7 SignedData  (primary — what Azure returns)
      2. base64 → bare DER X.509
      3. base64 → PEM text
      4. raw string → PEM text  (no base64 wrapping)

    Returns (None, None, None) if all attempts fail; a WARNING is logged with
    a data preview and each error so the cause can be diagnosed.
    """
    cert = None
    raw  = None
    err_pkcs7 = err_der = err_pem_raw = err_pem_str = None

    # Decode the base64 payload (restore any stripped padding first)
    try:
        padded = public_cert_data_b64.strip()
        padded += "=" * (-len(padded) % 4)
        raw = base64.b64decode(padded)
    except Exception as e:
        err_der = e  # record as DER error; no raw bytes to try further

    if raw is not None:
        # Attempt 1: PKCS#7 SignedData — the format Azure Application Gateway uses
        try:
            chain = _pkcs7.load_der_pkcs7_certificates(raw)
            if chain:
                cert = _leaf_cert(chain)
        except Exception as e:
            err_pkcs7 = e

        # Attempt 2: bare DER-encoded X.509 certificate
        if cert is None:
            try:
                cert = x509.load_der_x509_certificate(raw)
            except Exception as e:
                err_der = e

        # Attempt 3: PEM text that was itself base64-encoded
        if cert is None:
            try:
                cert = x509.load_pem_x509_certificate(raw)
            except Exception as e:
                err_pem_raw = e

    # Attempt 4: the value is already a PEM string (no base64 wrapping)
    if cert is None:
        try:
            cert = x509.load_pem_x509_certificate(public_cert_data_b64.encode())
        except Exception as e:
            err_pem_str = e

    if cert is None:
        preview = public_cert_data_b64[:120].replace("\n", "\\n") if public_cert_data_b64 else "<empty>"
        logger.warning(
            "Could not parse publicCertData — expiry will be unknown.\n"
            "  data preview : %s\n"
            "  err PKCS#7   : %s\n"
            "  err DER      : %s\n"
            "  err PEM(raw) : %s\n"
            "  err PEM(str) : %s",
            preview, err_pkcs7, err_der, err_pem_raw, err_pem_str,
        )
        return None, None, None

    subject = cert.subject.rfc4514_string()
    issuer  = cert.issuer.rfc4514_string()

    # not_valid_after_utc is timezone-aware (cryptography >= 42).
    # Older versions expose not_valid_after as a naive UTC datetime.
    try:
        expiry = cert.not_valid_after_utc
    except AttributeError:
        expiry = cert.not_valid_after.replace(tzinfo=datetime.timezone.utc)

    return subject, issuer, expiry


def _derive_status(days: Optional[int]) -> str:
    """Map days-remaining to a status label."""
    if days is None: return "Unknown"
    if days < 0:     return "Expired"
    if days <= 7:    return "Critical"
    if days <= 14:   return "Warning"
    return "OK"


def _get_subscription_name(credential, subscription_id: str) -> str:
    """
    Resolve a subscription's display name via the Subscription API.
    Falls back to the raw subscription ID if the lookup fails.
    """
    try:
        client = SubscriptionClient(credential)
        sub    = client.subscriptions.get(subscription_id)
        return sub.display_name or subscription_id
    except Exception as exc:
        logger.warning("Could not resolve name for subscription %s: %s", subscription_id, exc)
        return subscription_id


def _process_gateway(gw, subscription_id: str, subscription_name: str, now: datetime.datetime) -> List[dict]:
    """
    Extract one record per SSL certificate attached to an Application Gateway.

    Builds a cert → listener mapping in memory, then iterates ssl_certificates.
    Returns a list of flat dicts ready for Excel / HTML rendering.
    """
    rg = gw.id.split("/resourceGroups/")[1].split("/")[0]

    # Map each certificate name to the TLS listener(s) that reference it
    cert_to_listeners: Dict[str, List[str]] = {}
    for listener in (gw.http_listeners or []):
        if listener.ssl_certificate and listener.ssl_certificate.id:
            cert_name = listener.ssl_certificate.id.split("/")[-1]
            cert_to_listeners.setdefault(cert_name, []).append(listener.name)

    records: List[dict] = []
    ssl_certs = gw.ssl_certificates or []

    if not ssl_certs:
        logger.debug("Gateway %s has no SSL certificates.", gw.name)
        return records

    for cert in ssl_certs:
        is_kv   = bool(getattr(cert, "key_vault_secret_id", None))
        subject: Optional[str]           = None
        issuer:  Optional[str]           = None
        expiry:  Optional[datetime.datetime] = None

        if cert.public_cert_data:
            # Parse the PKCS#7 / DER / PEM blob returned by the Azure API
            subject, issuer, expiry = _parse_expiry(cert.public_cert_data)
        elif is_kv:
            # Key Vault-backed certs: Azure does not expose the raw cert data
            subject = "Key Vault reference (not exposed by API)"
            issuer  = "Key Vault reference"
            logger.warning(
                "Gateway %s / cert %s is Key Vault-backed — expiry unavailable.",
                gw.name, cert.name,
            )

        listeners = cert_to_listeners.get(cert.name, [])

        # Skip certificates not attached to any TLS listener — they are
        # unused / orphaned and do not affect live traffic.
        if not listeners:
            logger.debug(
                "Gateway %s / cert %s has no associated listeners — skipped.",
                gw.name, cert.name,
            )
            continue

        days   = (expiry - now).days if expiry else None
        status = _derive_status(days)

        records.append({
            "gateway_name":      gw.name,
            "gateway_location":  gw.location or "unknown",
            "resource_group":    rg,
            "subscription_name": subscription_name,
            "subscription_id":   subscription_id,
            "cert_name":         cert.name,
            "listeners":        listeners,
            "subject":          subject or "N/A",
            "issuer":           issuer  or "N/A",
            "expiry_date":      expiry,
            "days_remaining":   days,
            "status":           status,
            "key_vault_ref":    is_kv,
        })

    return records


def _scan_one_subscription(
    credential, subscription_id: str, now: datetime.datetime
) -> List[dict]:
    """
    Scan every Application Gateway in one subscription.

    list_all() is used first to discover gateways cheaply, then get() is called
    per gateway because list operations omit the publicCertData field.
    """
    client  = NetworkManagementClient(credential, subscription_id)
    sub_name = _get_subscription_name(credential, subscription_id)
    gw_refs = list(client.application_gateways.list_all())
    logger.info("[scan] subscription=%s (%s)  gateways found=%d", subscription_id, sub_name, len(gw_refs))

    records: List[dict] = []
    for gw_ref in gw_refs:
        try:
            # Individual GET returns full publicCertData that list_all() omits
            rg = gw_ref.id.split("/resourceGroups/")[1].split("/")[0]
            gw = client.application_gateways.get(
                resource_group_name=rg,
                application_gateway_name=gw_ref.name,
            )
            logger.info("[scan] Processing gateway %s / %s", rg, gw.name)
            records.extend(_process_gateway(gw, subscription_id, sub_name, now))
        except Exception as exc:
            logger.error("[scan] Failed to process gateway %s: %s", gw_ref.name, exc)

    return records


def scan_subscriptions(
    credential,
    subscription_ids: List[str],
    alert_days: int,
) -> List[dict]:
    """
    Enumerate ALL Application Gateway SSL certificates across subscriptions.

    Returns a flat list of dicts — one entry per (gateway, certificate) pair:
        gateway_name      str
        gateway_location  str
        resource_group    str
        subscription_name str
        subscription_id   str
        cert_name         str
        listeners        list[str]  — TLS listener names referencing this cert
        subject          str
        issuer           str
        expiry_date      datetime | None
        days_remaining   int | None
        status           str  — "OK" | "Warning" | "Critical" | "Expired" | "Unknown"
        key_vault_ref    bool
    """
    now         = datetime.datetime.now(datetime.timezone.utc)
    all_records: List[dict] = []

    for sub_id in subscription_ids:
        try:
            records = _scan_one_subscription(credential, sub_id, now)
            logger.info("[scan] subscription=%s  certs found=%d", sub_id, len(records))
            all_records.extend(records)
        except Exception as exc:
            logger.error("[scan] Failed to scan subscription %s: %s", sub_id, exc)

    return all_records


# =============================================================================
# SECTION 3 — Excel Report
# =============================================================================

# Column header colour (dark Azure blue) and per-status row fill colours
_FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
_FILL_EXPIRED = PatternFill("solid", fgColor="CC0000")
_FILL_CRIT    = PatternFill("solid", fgColor="FF4444")
_FILL_WARN    = PatternFill("solid", fgColor="FFA500")
_FILL_OK      = PatternFill("solid", fgColor="C6EFCE")
_FILL_UNKNOWN = PatternFill("solid", fgColor="DDDDDD")
_FONT_HEADER  = Font(color="FFFFFF", bold=True)

_EXCEL_COLUMNS: List[tuple] = [
    ("Gateway Name",       24),
    ("Location",           16),
    ("Resource Group",     22),
    ("Subscription Name",  28),
    ("Subscription ID",    36),
    ("Certificate Name",   26),
    ("Listeners",         30),
    ("Subject",           50),
    ("Issuer",            40),
    ("Expiry Date (UTC)", 22),
    ("Days Remaining",    16),
    ("Status",            12),
    ("Key Vault Ref",     14),
]


def _status_fill(status: str) -> PatternFill:
    return {
        "Expired":  _FILL_EXPIRED,
        "Critical": _FILL_CRIT,
        "Warning":  _FILL_WARN,
        "OK":       _FILL_OK,
    }.get(status, _FILL_UNKNOWN)


def build_excel(records: List[Dict[str, Any]]) -> bytes:
    """
    Build an .xlsx workbook covering every gateway / certificate record.

    Rows are sorted by days_remaining ascending (soonest expiry first);
    unknown-expiry rows appear at the bottom.  Each row is colour-coded by
    status.  Returns raw bytes suitable for email attachment.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title       = "SSL Certificates"
    ws.freeze_panes = "A2"  # keep the header row visible while scrolling

    # Write header row
    for col_idx, (label, width) in enumerate(_EXCEL_COLUMNS, 1):
        cell           = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = _FILL_HEADER
        cell.font      = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    # Sort: known expiry ascending, then unknown expiry rows
    sorted_records = sorted(
        records,
        key=lambda r: (r["days_remaining"] is None, r["days_remaining"] or 0),
    )

    for row_idx, rec in enumerate(sorted_records, 2):
        expiry_str    = rec["expiry_date"].strftime("%Y-%m-%d %H:%M") if rec["expiry_date"] else "N/A"
        listeners_str = ", ".join(rec["listeners"]) if rec["listeners"] else "—"
        days_val: Any = rec["days_remaining"] if rec["days_remaining"] is not None else "N/A"

        row_values = [
            rec["gateway_name"], rec["gateway_location"], rec["resource_group"],
            rec["subscription_name"], rec["subscription_id"], rec["cert_name"], listeners_str,
            rec["subject"], rec["issuer"], expiry_str, days_val,
            rec["status"], "Yes" if rec["key_vault_ref"] else "No",
        ]
        fill = _status_fill(rec["status"])
        for col_idx, value in enumerate(row_values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    ws.auto_filter.ref = ws.dimensions  # enable column filters

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# SECTION 4 — HTML Email
# =============================================================================

# Row background / foreground colours per status (used in both tables)
_ROW_BG = {"Expired": "#ffcccc", "Critical": "#ffcccc", "Warning": "#ffe0b2",
            "OK": "#F0FDF4", "Unknown": "#F8FAFC"}
_ROW_FG = {"Expired": "#7F1D1D", "Critical": "#7F1D1D", "Warning": "#78350F",
            "OK": "#14532D", "Unknown": "#475467"}


def _wrap_html_email(*, subject: str, body_html: str) -> str:
    """
    Wrap body_html in the branded email shell.

    Header gradient : Azure blue  (#1240A8 → #2563EB → #3B82F6)
    Section label   : "Azure Infrastructure"
    Title           : "SSL Certificate Expiry Monitor"
    Footer quote    : Swami Vivekananda
    """
    subj   = _html.escape(subject or "", quote=False)
    notice = _html.escape(
        "This is an automatically generated email. For questions or concerns, "
        "please contact the DevOps / Infrastructure team.",
        quote=False,
    )

    return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="color-scheme" content="light" />
    <meta name="supported-color-schemes" content="light" />
    <title>{subj}</title>
  </head>
  <body style="margin:0;padding:0;background-color:#EEF2F6;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0"
      style="border-collapse:collapse;background-color:#EEF2F6;">
      <tr>
        <td align="center" style="padding:32px 16px;">
          <table role="presentation" width="680" cellspacing="0" cellpadding="0" border="0"
            style="border-collapse:collapse;max-width:680px;width:100%;">
            <tr>
              <td style="padding:0;">
                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0"
                  style="border-collapse:collapse;background-color:#FFFFFF;border:1px solid #E4E7EC;
                  border-radius:16px;overflow:hidden;box-shadow:0 8px 32px rgba(15,23,42,0.08);">

                  <!-- Header banner -->
                  <tr>
                    <td style="padding:0;background-color:#175CD3;
                      background-image:linear-gradient(135deg,#1240A8 0%,#2563EB 55%,#3B82F6 100%);">
                      <div style="padding:22px 24px;font-family:Segoe UI,Roboto,Arial,sans-serif;">
                        <div style="font-size:11px;font-weight:600;letter-spacing:0.12em;
                          text-transform:uppercase;color:#BFDBFE;margin-bottom:8px;">
                          DevOps Team - Automated Notifier
                        </div>
                        <div style="font-size:20px;font-weight:700;line-height:1.25;color:#FFFFFF;
                          letter-spacing:-0.02em;">
                          SSL Certificate Expiry Monitor
                        </div>
                        <div style="margin-top:8px;font-size:13px;color:#DBEAFE;line-height:1.45;">
                          Daily Application Gateway certificate health check
                        </div>
                      </div>
                    </td>
                  </tr>

                  <!-- Subject line -->
                  <tr>
                    <td style="padding:22px 24px 8px 24px;background-color:#FFFFFF;">
                      <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:17px;
                        font-weight:700;line-height:1.35;color:#101828;letter-spacing:-0.02em;">
                        {subj}
                      </div>
                      <div style="margin-top:8px;height:3px;width:48px;border-radius:2px;
                        background:linear-gradient(90deg,#175CD3,#60A5FA);"></div>
                    </td>
                  </tr>

                  <!-- Body content -->
                  <tr>
                    <td style="padding:8px 24px 24px 24px;background-color:#FFFFFF;">
                      <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:14px;
                        line-height:1.6;color:#344054;">
                        {body_html}
                      </div>
                    </td>
                  </tr>

                  <!-- Closing card -->
                  <tr>
                    <td style="padding:0 24px 20px 24px;background-color:#FFFFFF;">
                      <div style="padding:16px 18px;border-radius:12px;background-color:#F0FDF4;
                        border:1px solid #BBF7D0;text-align:center;">
                        <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:16px;
                          font-weight:600;color:#166534;line-height:1.4;">
                          Have a wonderful day
                        </div>
                        <div style="margin-top:6px;font-family:Segoe UI,Roboto,Arial,sans-serif;
                          font-size:13px;color:#15803D;line-height:1.5;">
                          Arise, awake, and stop not till the goal is reached.
                          — <strong>Swami Vivekananda</strong>
                        </div>
                      </div>
                    </td>
                  </tr>

                  <!-- Footer -->
                  <tr>
                    <td style="padding:18px 24px;background-color:#F9FAFB;
                      border-top:1px solid #E5E7EB;">
                      <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:12px;
                        line-height:1.55;color:#667085;">
                        <span style="display:inline-block;padding:2px 8px;border-radius:6px;
                          background-color:#EFF4FF;color:#175CD3;font-weight:600;font-size:11px;
                          margin-right:6px;">Auto</span>
                        <strong style="color:#475467;">Notice:</strong> {notice}
                      </div>
                    </td>
                  </tr>

                </table>

                <div style="margin-top:16px;font-family:Segoe UI,Roboto,Arial,sans-serif;
                  font-size:11px;color:#98A2B3;text-align:center;line-height:1.5;">
                  This message was sent automatically. Please do not reply to this email.
                </div>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""


def _status_cert_table_html(
    records: List[Dict[str, Any]],
    title: str,
    subtitle: str,
    header_gradient: str,
    col_header_bg: str,
) -> str:
    """
    Render a certificate table for a single status group (Expired / Critical / Warning).

    Args:
        records:          Filtered + sorted list of cert records for this status.
        title:            Table header title  (e.g. "Top 5 — Expired Certificates").
        subtitle:         Table header subtitle.
        header_gradient:  CSS background-image gradient for the title row.
        col_header_bg:    Background colour of the column-label row.

    Returns an empty string if records is empty (table is omitted from the email).
    """
    if not records:
        return ""

    td = (
        "border-bottom:1px solid #E8ECF2;"
        "font-family:Segoe UI,Roboto,Arial,sans-serif;"
        "vertical-align:middle;"
    )
    th = (
        "padding:12px 12px;font-size:10px;font-weight:800;text-transform:uppercase;"
        "letter-spacing:0.08em;color:#FFFFFF;font-family:Segoe UI,Roboto,Arial,sans-serif;"
    )

    badge_bg = {"Expired": "#FEE2E2", "Critical": "#FEE2E2", "Warning": "#FEF3C7"}.get(records[0]["status"], "#F1F5F9")
    badge_fg = {"Expired": "#991B1B", "Critical": "#991B1B", "Warning": "#92400E"}.get(records[0]["status"], "#475467")

    rows: List[str] = []
    for i, r in enumerate(records):
        bg      = _ROW_BG.get(r["status"], "#FFFFFF")
        last    = "border-bottom:none;" if i == len(records) - 1 else ""
        expiry  = r["expiry_date"].strftime("%Y-%m-%d") if r["expiry_date"] else "N/A"
        listen  = _html.escape(", ".join(r["listeners"]) if r["listeners"] else "—")
        status  = _html.escape(r["status"])
        days    = r["days_remaining"]
        days_s  = f"{days}d" if days is not None else "N/A"

        rows.append(f"""
        <tr style="background-color:{bg};">
          <td style="padding:14px 14px 14px 18px;{td}{last}
            font-size:14px;color:#475467;text-align:center;font-weight:800;width:44px;">{i+1}</td>
          <td style="padding:14px 12px;{td}{last}
            font-size:14px;color:#101828;font-weight:600;line-height:1.4;">
            {_html.escape(r['gateway_name'])}</td>
          <td style="padding:14px 12px;{td}{last}font-size:13px;color:#344054;">
            {_html.escape(r['resource_group'])}</td>
          <td style="padding:14px 12px;{td}{last}font-size:12px;color:#667085;">
            {_html.escape(r['subscription_name'])}</td>
          <td style="padding:14px 12px;{td}{last}
            font-size:12px;color:#667085;font-family:ui-monospace,Consolas,monospace;">
            {_html.escape(r['cert_name'])}</td>
          <td style="padding:14px 12px;{td}{last}font-size:12px;color:#667085;">{listen}</td>
          <td style="padding:14px 12px;{td}{last}font-size:13px;color:#101828;white-space:nowrap;">
            {_html.escape(expiry)}</td>
          <td style="padding:14px 20px 14px 12px;{td}{last}text-align:center;">
            <span style="display:inline-block;padding:4px 10px;border-radius:20px;
              font-size:13px;font-weight:800;font-family:Segoe UI,Roboto,Arial,sans-serif;
              background-color:{badge_bg};color:{badge_fg};letter-spacing:-0.01em;">
              {days_s} &nbsp;·&nbsp; {status}
            </span>
          </td>
        </tr>""")

    return f"""
<div style="margin:22px 0 0 0;border-radius:14px;overflow:hidden;border:1px solid #D8DEE6;
  box-shadow:0 6px 20px rgba(15,23,42,0.07);">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0"
    style="border-collapse:collapse;">
    <thead>
      <tr>
        <th colspan="8" align="left"
          style="padding:18px 20px;{header_gradient}">
          <span style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:14px;
            font-weight:700;color:#FFFFFF;letter-spacing:-0.02em;">{_html.escape(title)}</span>
          <span style="display:block;margin-top:6px;font-size:12px;font-weight:400;
            color:rgba(255,255,255,0.75);font-family:Segoe UI,Roboto,Arial,sans-serif;
            line-height:1.45;">{_html.escape(subtitle)}</span>
        </th>
      </tr>
      <tr style="background-color:{col_header_bg};">
        <th align="center" style="{th}width:44px;padding-left:18px;">#</th>
        <th align="left"   style="{th}">App Gateway</th>
        <th align="left"   style="{th}">Resource Group</th>
        <th align="left"   style="{th}">Subscription</th>
        <th align="left"   style="{th}">Certificate</th>
        <th align="left"   style="{th}">Listener(s)</th>
        <th align="left"   style="{th}">Expiry Date</th>
        <th align="center" style="{th}padding-right:20px;">Days / Status</th>
      </tr>
    </thead>
    <tbody>{"".join(rows)}</tbody>
  </table>
</div>""".strip()


def _unknown_cert_table_html(unknown: List[Dict[str, Any]]) -> str:
    """
    Render all certs with unknown expiry as a separate grey-themed table.
    Returns an empty string if there are no unknown-expiry certs.
    """
    if not unknown:
        return ""

    td = (
        "border-bottom:1px solid #E8ECF2;"
        "font-family:Segoe UI,Roboto,Arial,sans-serif;"
        "vertical-align:middle;"
    )
    th = (
        "padding:11px 12px;font-size:10px;font-weight:800;text-transform:uppercase;"
        "letter-spacing:0.08em;color:#FFFFFF;font-family:Segoe UI,Roboto,Arial,sans-serif;"
    )

    rows: List[str] = []
    for i, r in enumerate(unknown):
        last   = "border-bottom:none;" if i == len(unknown) - 1 else ""
        listen = _html.escape(", ".join(r["listeners"]) if r["listeners"] else "—")
        # Distinguish between Key Vault certs (Azure limitation) and parse failures
        reason = "Key Vault reference" if r["key_vault_ref"] else "Expiry unreadable"

        rows.append(f"""
        <tr style="background-color:#F8FAFC;">
          <td style="padding:12px 14px 12px 18px;{td}{last}
            font-size:13px;color:#667085;text-align:center;font-weight:700;width:44px;">{i+1}</td>
          <td style="padding:12px 12px;{td}{last}font-size:13px;color:#101828;font-weight:600;">
            {_html.escape(r['gateway_name'])}</td>
          <td style="padding:12px 12px;{td}{last}font-size:13px;color:#344054;">
            {_html.escape(r['resource_group'])}</td>
          <td style="padding:12px 12px;{td}{last}font-size:12px;color:#667085;">
            {_html.escape(r['subscription_name'])}</td>
          <td style="padding:12px 12px;{td}{last}
            font-size:12px;color:#667085;font-family:ui-monospace,Consolas,monospace;">
            {_html.escape(r['cert_name'])}</td>
          <td style="padding:12px 12px;{td}{last}font-size:12px;color:#667085;">{listen}</td>
          <td style="padding:12px 20px 12px 12px;{td}{last}">
            <span style="display:inline-block;padding:3px 10px;border-radius:20px;
              font-size:12px;font-weight:700;background-color:#F1F5F9;color:#475467;">
              {_html.escape(reason)}
            </span>
          </td>
        </tr>""")

    count = len(unknown)
    return f"""
<div style="margin:22px 0 0 0;border-radius:14px;overflow:hidden;border:1px solid #D8DEE6;
  box-shadow:0 4px 14px rgba(15,23,42,0.06);">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0"
    style="border-collapse:collapse;">
    <thead>
      <tr>
        <th colspan="7" align="left"
          style="padding:16px 20px;background-color:#475467;
          background-image:linear-gradient(135deg,#344054 0%,#667085 100%);">
          <span style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:14px;
            font-weight:700;color:#FFFFFF;letter-spacing:-0.02em;">
            Unknown Expiry — {count} Certificate{'s' if count != 1 else ''}
          </span>
          <span style="display:block;margin-top:5px;font-size:12px;font-weight:400;
            color:#D0D5DD;font-family:Segoe UI,Roboto,Arial,sans-serif;line-height:1.45;">
            Expiry date unavailable — verify these certificates manually
          </span>
        </th>
      </tr>
      <tr style="background-color:#667085;">
        <th align="center" style="{th}width:44px;padding-left:18px;">#</th>
        <th align="left"   style="{th}">App Gateway</th>
        <th align="left"   style="{th}">Resource Group</th>
        <th align="left"   style="{th}">Subscription</th>
        <th align="left"   style="{th}">Certificate</th>
        <th align="left"   style="{th}">Listener(s)</th>
        <th align="left"   style="{th}padding-right:20px;">Reason</th>
      </tr>
    </thead>
    <tbody>{"".join(rows)}</tbody>
  </table>
</div>""".strip()


def build_html_top10(records: List[Dict[str, Any]]) -> str:
    """
    Build the full branded HTML email document.

    Contains:
      - At-a-glance metrics card (counts per status)
      - Top-5 Expired certificates table   (dark-red theme)
      - Top-5 Critical certificates table  (orange-red theme)
      - Top-5 Warning certificates table   (amber theme)
      - Unknown-expiry certificates table  (grey theme, if any)
      - Legend
    """
    now_str = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    unknown = [r for r in records if r["days_remaining"] is None]

    # Counts per status
    n_expired  = sum(1 for r in records if r["status"] == "Expired")
    n_critical = sum(1 for r in records if r["status"] == "Critical")
    n_warning  = sum(1 for r in records if r["status"] == "Warning")
    n_ok       = sum(1 for r in records if r["status"] == "OK")
    n_unknown  = len(unknown)
    n_total    = len(records)

    # Top-5 per status, sorted by days_remaining ascending (most urgent first)
    def _top5(status: str) -> List[Dict[str, Any]]:
        return sorted(
            [r for r in records if r["status"] == status],
            key=lambda r: r["days_remaining"],
        )[:5]

    expired_top5  = _top5("Expired")
    critical_top5 = _top5("Critical")
    warning_top5  = _top5("Warning")

    # Render the three status tables with distinct colour themes
    expired_html  = _status_cert_table_html(
        expired_top5,
        title    = f"Top 5 — Expired Certificates  ({n_expired} total)",
        subtitle = "These certificates have already expired and must be renewed immediately",
        header_gradient = "background-color:#991B1B;background-image:linear-gradient(135deg,#7F1D1D 0%,#DC2626 100%);",
        col_header_bg   = "#B91C1C",
    )
    critical_html = _status_cert_table_html(
        critical_top5,
        title    = f"Top 5 — Critical Certificates  ({n_critical} total)  ≤ 7 days",
        subtitle = "These certificates expire within 7 days — renew urgently",
        header_gradient = "background-color:#9A3412;background-image:linear-gradient(135deg,#7C2D12 0%,#EA580C 100%);",
        col_header_bg   = "#C2410C",
    )
    warning_html  = _status_cert_table_html(
        warning_top5,
        title    = f"Top 5 — Warning Certificates  ({n_warning} total)  ≤ 14 days",
        subtitle = "These certificates expire within 14 days — schedule renewal soon",
        header_gradient = "background-color:#92400E;background-image:linear-gradient(135deg,#78350F 0%,#D97706 100%);",
        col_header_bg   = "#B45309",
    )
    unknown_html  = _unknown_cert_table_html(unknown)

    # Alert badges — only shown for non-zero counts
    badge_parts: List[str] = []
    for count, label, bg, fg in [
        (n_expired,  "Expired",            "#FEE2E2", "#991B1B"),
        (n_critical, "Critical (≤7 days)", "#FEE2E2", "#991B1B"),
        (n_warning,  "Warning (≤14 days)", "#FEF3C7", "#92400E"),
    ]:
        if count:
            badge_parts.append(
                f'<span style="display:inline-block;padding:3px 10px;border-radius:20px;'
                f'background-color:{bg};color:{fg};font-weight:700;font-size:13px;'
                f'font-family:Segoe UI,Roboto,Arial,sans-serif;margin:2px 4px 2px 0;">'
                f'{count} {_html.escape(label)}</span>'
            )
    alert_badges = (
        " ".join(badge_parts)
        if badge_parts
        else '<span style="color:#166534;font-weight:600;">All certificates are healthy ✓</span>'
    )

    ml = (  # metric label style (reused across the grid)
        "font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:0.07em;"
        "color:#667085;font-family:Segoe UI,Roboto,Arial,sans-serif;margin-bottom:8px;"
    )

    body_html = f"""
<p style="margin:0 0 16px 0;font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:16px;
  font-weight:600;color:#101828;letter-spacing:-0.01em;">Hi Team,</p>

<div style="margin:0 0 18px 0;padding:18px 20px;border-radius:14px;background-color:#F8FAFF;
  border:1px solid #E0E7FF;border-left:4px solid #175CD3;">
  <p style="margin:0;font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:15px;
    line-height:1.65;color:#1D2939;">
    Please find the daily SSL certificate health report below. Certificates are grouped by
    status — <strong>Expired</strong>, <strong>Critical (≤7 days)</strong>, and
    <strong>Warning (≤14 days)</strong> — showing the top 5 most urgent in each group.
    The full report for <strong>all Application Gateway certificates</strong> is attached
    as an Excel file.
  </p>
</div>

<!-- At a glance metrics -->
<div style="margin:0 0 20px 0;padding:18px 20px;border-radius:14px;background-color:#FCFCFD;
  border:1px solid #E4E7EC;">
  <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:11px;font-weight:800;
    text-transform:uppercase;letter-spacing:0.06em;color:#98A2B3;margin-bottom:14px;">At a glance</div>
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0"
    style="border-collapse:collapse;">
    <tr>
      <td width="50%" valign="top"
        style="padding:14px 16px 14px 0;border-bottom:1px solid #EEF2F6;">
        <div style="{ml}">Report generated (UTC)</div>
        <div style="font-size:14px;font-weight:700;color:#175CD3;
          font-family:Segoe UI,Roboto,Arial,sans-serif;">{_html.escape(now_str)}</div>
      </td>
      <td width="50%" valign="top"
        style="padding:14px 0 14px 16px;border-bottom:1px solid #EEF2F6;
        border-left:1px solid #E4E7EC;">
        <div style="{ml}">Total certificates scanned</div>
        <div style="font-size:22px;font-weight:800;color:#101828;
          font-family:Segoe UI,Roboto,Arial,sans-serif;letter-spacing:-0.03em;">{n_total}</div>
      </td>
    </tr>
    <tr>
      <td width="50%" valign="top" style="padding:14px 16px 0 0;">
        <div style="{ml}">Alert status</div>
        <div style="margin-top:4px;">{alert_badges}</div>
      </td>
      <td width="50%" valign="top"
        style="padding:14px 0 0 16px;border-left:1px solid #E4E7EC;">
        <div style="{ml}">Healthy (OK) / Unknown expiry</div>
        <div style="font-size:20px;font-weight:800;color:#166534;
          font-family:Segoe UI,Roboto,Arial,sans-serif;">{n_ok}
          <span style="font-size:14px;font-weight:600;color:#667085;margin-left:6px;">
            / {n_unknown} unknown
          </span>
        </div>
      </td>
    </tr>
  </table>
</div>

{expired_html}

{critical_html}

{warning_html}

{unknown_html}

<!-- Legend -->
<div style="margin:20px 0 0 0;padding:14px 16px;border-radius:10px;background-color:#F8FAFC;
  border:1px solid #E2E8F0;">
  <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:11px;font-weight:800;
    text-transform:uppercase;letter-spacing:0.07em;color:#94A3B8;margin-bottom:10px;">Legend</div>
  <div style="font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:13px;">
    <span style="display:inline-block;padding:3px 10px;border-radius:20px;
      background-color:#FEE2E2;color:#991B1B;font-weight:700;margin:2px 6px 2px 0;">
      Expired / Critical &le;7 days</span>
    <span style="display:inline-block;padding:3px 10px;border-radius:20px;
      background-color:#FEF3C7;color:#92400E;font-weight:700;margin:2px 6px 2px 0;">
      Warning &le;14 days</span>
    <span style="display:inline-block;padding:3px 10px;border-radius:20px;
      background-color:#DCFCE7;color:#166534;font-weight:700;margin:2px 6px 2px 0;">
      OK</span>
  </div>
</div>

<p style="margin:20px 0 0 0;font-family:Segoe UI,Roboto,Arial,sans-serif;font-size:13px;
  color:#667085;line-height:1.6;">
  Full details for all Application Gateway certificates — including subscription, location,
  issuer, subject, and listener mappings — are in the attached Excel file.
</p>""".strip()

    subject_for_wrapper = (
        f"SSL Certificate Expiry Report — "
        f"{n_expired + n_critical} critical, {n_warning} warning, {n_total} total"
    )
    return _wrap_html_email(subject=subject_for_wrapper, body_html=body_html)


# =============================================================================
# SECTION 5 — ACS Email Notification
# =============================================================================

def _send_via_acs(
    connection_string: str,
    sender: str,
    to: str,
    subject: str,
    html_body: str,
    excel_filename: str,
    excel_bytes: bytes,
) -> None:
    """
    Send the notification email via Azure Communication Services (ACS) Email.

    Recipients are derived from the comma-separated *to* string.
    The Excel report is attached as a base64-encoded inline attachment.

    Raises an exception on failure so the caller can log and exit.
    """
    from azure.communication.email import EmailClient  # type: ignore

    recipients = [addr.strip() for addr in to.split(",") if addr.strip()]

    message = {
        "senderAddress": sender,
        "recipients": {
            "to": [{"address": addr} for addr in recipients],
        },
        "content": {
            "subject": subject,
            "html": html_body,
        },
        "attachments": [
            {
                "name": excel_filename,
                "contentType": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                "contentInBase64": base64.b64encode(excel_bytes).decode("ascii"),
            }
        ],
    }

    logger.info("[notify] Sending via ACS Email — to=%r  attachment=%r", to, excel_filename)
    client  = EmailClient.from_connection_string(connection_string)
    poller  = client.begin_send(message)
    result  = poller.result()
    logger.info("[notify] ACS Email sent — message_id=%s", result.get("id", "N/A"))


# =============================================================================
# SECTION 6 — Entry Point
# =============================================================================

def main() -> None:
    cfg = _load_config()

    subscription_ids = [s.strip() for s in cfg["subscription_ids"].split(",") if s.strip()]
    alert_days       = cfg["alert_days"]

    logger.info(
        "[main] Starting scan — subscriptions=%d  alert_days=%d  dry_run=%s",
        len(subscription_ids), alert_days, cfg["dry_run"],
    )

    # Authenticate with Azure using a service principal (read-only operations only)
    credential = ClientSecretCredential(
        tenant_id=cfg["tenant_id"],
        client_id=cfg["client_id"],
        client_secret=cfg["client_secret"],
    )

    # Scan every subscription — list gateways, GET each for full cert data
    all_records = scan_subscriptions(credential, subscription_ids, alert_days)
    logger.info("[main] Total certificates found: %d", len(all_records))

    if not all_records:
        logger.info("[main] No certificates found. Nothing to report.")
        return

    today = datetime.date.today().strftime("%Y-%m-%d")

    # Build Excel attachment (all certs, colour-coded by status)
    excel_bytes    = build_excel(all_records)
    excel_filename = f"ssl_certificate_report_{today}.xlsx"
    logger.info("[main] Excel report built — %d rows  file=%s", len(all_records), excel_filename)

    # Build HTML email body (top-10 soonest expiring + unknown-expiry table)
    html_body = build_html_top10(all_records)

    # Compose email subject with alert counts
    n_crit  = sum(1 for r in all_records if r["status"] in ("Expired", "Critical"))
    n_warn  = sum(1 for r in all_records if r["status"] == "Warning")
    subject = (
        f"[SSL Alert] App Gateway Cert Expiry — "
        f"{n_crit} critical, {n_warn} warning, {len(all_records)} total — {today}"
    )

    if cfg["dry_run"]:
        # Dry-run: write outputs locally instead of sending the email
        with open(excel_filename, "wb") as f:
            f.write(excel_bytes)
        html_path = f"dry_run_email_{today}.html"
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_body)
        logger.info("[main] DRY RUN — Excel saved to  : %s", excel_filename)
        logger.info("[main] DRY RUN — Email HTML saved : %s", html_path)
        logger.info("[main] DRY RUN — Subject would be : %s", subject)
        logger.info("[main] DRY RUN — No ACS Email call made.")
    else:
        _send_via_acs(
            connection_string=cfg["acs_connection_string"],
            sender=cfg["acs_sender"],
            to=cfg["notify_to"],
            subject=subject,
            html_body=html_body,
            excel_filename=excel_filename,
            excel_bytes=excel_bytes,
        )
        logger.info("[main] Done. Notification sent.")


if __name__ == "__main__":
    main()
